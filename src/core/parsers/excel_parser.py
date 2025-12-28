"""Excel document parser for table schema extraction."""

from pathlib import Path
from typing import Any, Optional
import io

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.core.parsers.base import (
    BaseParser,
    ContentType,
    ParsedDocument,
    ParsedSection,
)


class ExcelParser(BaseParser):
    """Parser for Excel (.xlsx) files, optimized for table schema extraction."""

    SUPPORTED_EXTENSIONS = {".xlsx", ".xls"}

    # Common column name patterns for schema detection
    SCHEMA_COLUMN_PATTERNS = {
        "column", "field", "name", "attribute",
        "type", "datatype", "data_type",
        "description", "desc", "comment", "note",
        "nullable", "null", "required", "optional",
        "default", "default_value",
        "constraint", "key", "pk", "fk", "primary", "foreign",
        "length", "size", "precision", "scale",
    }

    def parse(self, file_path: Path) -> ParsedDocument:
        """Parse an Excel file."""
        wb = load_workbook(file_path, read_only=True, data_only=True)
        return self._parse_workbook(wb, file_path.name, file_path)

    def parse_bytes(self, content: bytes, filename: str) -> ParsedDocument:
        """Parse Excel from bytes."""
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        return self._parse_workbook(wb, filename, Path(filename))

    def _parse_workbook(
        self,
        wb: Any,
        filename: str,
        file_path: Path,
    ) -> ParsedDocument:
        """Parse Excel workbook into sections."""
        sections: list[ParsedSection] = []
        metadata: dict[str, Any] = {
            "sheet_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
            "tables": [],
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_sections = self._parse_sheet(ws, sheet_name)
            sections.extend(sheet_sections)

            # Collect table metadata
            if sheet_sections:
                for section in sheet_sections:
                    if section.metadata.get("is_schema_table"):
                        metadata["tables"].append({
                            "sheet": sheet_name,
                            "columns": section.metadata.get("columns", []),
                        })

        return ParsedDocument(
            filename=filename,
            file_path=file_path,
            sections=sections,
            metadata=metadata,
        )

    def _parse_sheet(
        self,
        ws: Worksheet,
        sheet_name: str,
    ) -> list[ParsedSection]:
        """Parse a single worksheet."""
        sections: list[ParsedSection] = []

        # Get all data
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return sections

        # Find header row (first non-empty row)
        header_row_idx = 0
        headers = []
        for idx, row in enumerate(rows):
            if any(cell is not None for cell in row):
                headers = [str(cell) if cell else f"col_{i}" for i, cell in enumerate(row)]
                header_row_idx = idx
                break

        if not headers:
            return sections

        # Check if this looks like a schema table
        is_schema = self._is_schema_table(headers)

        # Convert to Markdown table
        table_content = self._rows_to_markdown(headers, rows[header_row_idx + 1:])

        # Create section
        section = ParsedSection(
            content=table_content,
            content_type=ContentType.TABLE,
            heading_text=sheet_name,
            section_hierarchy=[sheet_name],
            metadata={
                "sheet_name": sheet_name,
                "is_schema_table": is_schema,
                "columns": headers,
                "row_count": len(rows) - header_row_idx - 1,
            },
        )

        sections.append(section)

        # If it's a schema table, also create text description
        if is_schema:
            description = self._generate_schema_description(headers, rows[header_row_idx + 1:])
            sections.append(
                ParsedSection(
                    content=description,
                    content_type=ContentType.TEXT,
                    heading_text=f"{sheet_name} (Schema Description)",
                    section_hierarchy=[sheet_name],
                    metadata={"is_generated_description": True},
                )
            )

        return sections

    def _is_schema_table(self, headers: list[str]) -> bool:
        """Check if headers suggest this is a database schema table."""
        header_lower = [h.lower() for h in headers]

        # Count matching schema patterns
        matches = sum(
            1 for h in header_lower
            for pattern in self.SCHEMA_COLUMN_PATTERNS
            if pattern in h
        )

        # If multiple schema-like columns, it's probably a schema table
        return matches >= 2

    def _rows_to_markdown(
        self,
        headers: list[str],
        data_rows: list[tuple],
    ) -> str:
        """Convert rows to Markdown table format."""
        lines = []

        # Header row
        header_line = "| " + " | ".join(headers) + " |"
        lines.append(header_line)

        # Separator
        separator = "| " + " | ".join(["---"] * len(headers)) + " |"
        lines.append(separator)

        # Data rows
        for row in data_rows:
            if any(cell is not None for cell in row):
                cells = [
                    str(cell).replace("|", "\\|").replace("\n", " ") if cell else ""
                    for cell in row[:len(headers)]
                ]
                # Pad if row is shorter than headers
                while len(cells) < len(headers):
                    cells.append("")
                line = "| " + " | ".join(cells) + " |"
                lines.append(line)

        return "\n".join(lines)

    def _generate_schema_description(
        self,
        headers: list[str],
        data_rows: list[tuple],
    ) -> str:
        """Generate a text description of the schema table."""
        lines = ["Table Schema:"]

        for row in data_rows[:20]:  # Limit to first 20 rows
            if not any(cell is not None for cell in row):
                continue

            parts = []
            for i, (header, cell) in enumerate(zip(headers, row)):
                if cell:
                    header_lower = header.lower()
                    if any(p in header_lower for p in ["name", "column", "field"]):
                        parts.insert(0, f"**{cell}**")
                    elif any(p in header_lower for p in ["type", "datatype"]):
                        parts.append(f"({cell})")
                    elif any(p in header_lower for p in ["desc", "comment", "note"]):
                        parts.append(f"- {cell}")

            if parts:
                lines.append(" ".join(parts))

        return "\n".join(lines)
